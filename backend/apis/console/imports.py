from __future__ import annotations

from typing import Annotated, Iterable
from decimal import Decimal
from io import BytesIO
import csv

import arrow
from fastapi import Header, UploadFile, File, Form, Query
from pydantic import BaseModel, Field
import chardet

from backend.apis import *
from backend.apis.auth.base import *
from backend.core.tpdm import *
from backend.services import Privileges
from backend.models import (
    Customer,
    MonthBill,
    MonthBillSettlementType,
    Order,
    OrderOrigin,
    OrderStatus,
    ProductType,
    WalletAccount,
    WalletTransaction,
    WalletTransactionType,
)


class ImportTextBody(BaseModel):
    text: str = Field(description="导入文本（支持 TSV：制表符分隔，首行表头）")
    skip_header: bool = Field(default=True, description="是否跳过首行表头")


class ImportResult(BaseModel):
    created: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0


def _lines(text: str) -> list[str]:
    return [i for i in (text or "").splitlines() if i and i.strip()]


def _normalize_cell(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _detect_encoding(data: bytes) -> str:
    if not data:
        return "utf-8"
    r = chardet.detect(data)
    enc = (r or {}).get("encoding") or "utf-8"
    return enc


def _iter_table_rows_from_csv_bytes(data: bytes) -> Iterable[list[str]]:
    enc = _detect_encoding(data)
    text = data.decode(enc, errors="ignore")
    # 兼容 Excel 导出的 UTF-8 BOM
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";"])
    except Exception:
        class _D:
            delimiter = "\t" if "\t" in sample else ","
        dialect = _D()
    reader = csv.reader(text.splitlines(), delimiter=getattr(dialect, "delimiter", ","))
    for row in reader:
        yield [_normalize_cell(i) for i in row]


def _iter_table_rows_from_xlsx_bytes(data: bytes) -> Iterable[list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise BadRequest(f"缺少 Excel 解析依赖 openpyxl: {e}")
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        yield [_normalize_cell(i) for i in row]


async def _iter_rows_from_upload(file: UploadFile) -> Iterable[list[str]]:
    filename = (file.filename or "").lower()
    data = await file.read()
    if filename.endswith(".csv"):
        return _iter_table_rows_from_csv_bytes(data)
    if filename.endswith(".xlsx"):
        return _iter_table_rows_from_xlsx_bytes(data)
    raise BadRequest("仅支持 .csv / .xlsx 文件")


async def _import_orders_from_rows(rows: Iterable[list[str]], skip_header: bool) -> ImportResult:
    result = ImportResult()
    idx = 0
    for cols in rows:
        if skip_header and idx == 0:
            idx += 1
            continue
        idx += 1
        cols = [c.strip() for c in cols if c is not None]
        # 订单导入：必须包含 customer_code，用于关联客户信息
        if len(cols) < 11:
            result.failed += 1
            continue

        (
            order_no,
            batch_code,
            product_name,
            resource_code,
            order_guanxi_id,
            customer_code,
            origin,
            product_type,
            order_date,
            status,
            total_price,
        ) = cols[:11]

        try:
            origin_v = OrderOrigin.get_by_name(origin).value
            product_type_v = ProductType.get_by_name(product_type).value
            status_v = OrderStatus.get_by_name(status).value
            order_date_v = arrow.get(order_date).date()
            total_price_v = Decimal(str(total_price))
        except Exception:
            result.failed += 1
            continue
        if not (customer_code or "").strip():
            result.failed += 1
            continue
        try:
            await _require_customer_by_code(customer_code)
        except Exception:
            result.failed += 1
            continue

        exist = await Order.filter(order_id=order_no).first()
        if exist:
            exist.customer_code = customer_code or exist.customer_code
            exist.batch_code = batch_code or exist.batch_code
            exist.product_name = product_name
            exist.resource_code = resource_code
            exist.order_guanxi_id = order_guanxi_id
            exist.origin = origin_v
            exist.product_type = product_type_v
            exist.order_date = order_date_v
            exist.status = status_v
            exist.total_price = total_price_v
            await exist.save()
            result.updated += 1
        else:
            await Order.create(
                order_id=order_no,
                customer_code=customer_code,
                batch_code=batch_code,
                product_name=product_name,
                resource_code=resource_code,
                order_guanxi_id=order_guanxi_id,
                origin=origin_v,
                product_type=product_type_v,
                order_date=order_date_v,
                status=status_v,
                total_price=total_price_v,
            )
            result.created += 1
    return result


async def _import_month_bills_from_rows(rows: Iterable[list[str]], skip_header: bool) -> ImportResult:
    result = ImportResult()
    idx = 0
    for cols in rows:
        if skip_header and idx == 0:
            idx += 1
            continue
        idx += 1
        cols = [c.strip() for c in cols if c is not None]
        if len(cols) < 12:
            result.failed += 1
            continue

        (
            yearmonth,
            product_name,
            customer_code,
            customer_name,
            account_code,
            product_type,
            resource_code,
            amount,
            settlement_type,
            homepage_price,
            discount_price,
            total_paid,
        ) = cols[:12]

        try:
            year = int(yearmonth[:4])
            month = int(yearmonth[-2:])
            product_type_v = ProductType.get_by_name(product_type).value
            settlement_type_v = MonthBillSettlementType.get_by_name(settlement_type).value
            amount_v = int(amount) if amount not in (None, "", "-") else None
            homepage_price_v = Decimal(str(homepage_price or 0))
            discount_price_v = Decimal(str(discount_price or 0))
            total_paid_v = Decimal(str(total_paid or 0))
        except Exception:
            result.failed += 1
            continue

        try:
            await _require_customer_by_code(customer_code)
        except Exception:
            result.failed += 1
            continue

        exist = await MonthBill.filter(
            year=year,
            month=month,
            customer_code=customer_code,
            account_code=account_code,
            product_type=product_type_v,
            resource_code=resource_code,
            settlement_type=settlement_type_v,
        ).first()

        if exist:
            exist.product_name = product_name
            exist.customer_name = customer_name
            exist.amount = amount_v
            exist.homepage_price = homepage_price_v
            exist.discount_price = discount_price_v
            exist.total_paid = total_paid_v
            await exist.save()
            result.updated += 1
        else:
            await MonthBill.create(
                year=year,
                month=month,
                product_name=product_name,
                customer_code=customer_code,
                customer_name=customer_name,
                account_code=account_code,
                product_type=product_type_v,
                resource_code=resource_code,
                amount=amount_v,
                settlement_type=settlement_type_v,
                homepage_price=homepage_price_v,
                discount_price=discount_price_v,
                total_paid=total_paid_v,
            )
            result.created += 1
    return result


def _parse_bool(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "on", "是")


def _normalize_cost_type(v: str) -> str:
    s = (v or "").strip().lower()
    if s in ("流量", "traffic", "stream"):
        return "stream"
    if s in ("短信", "sms", "短信服务"):
        return "sms"
    if s in ("token", "tokens"):
        return "token"
    return s or "unknown"


def _build_cost_remark(cost_type: str, remark: str | None) -> str | None:
    base = (remark or "").strip()
    ct = _normalize_cost_type(cost_type)
    prefix = {"stream": "[stream]", "sms": "[sms]", "token": "[token]"}.get(ct, f"[{ct}]")
    if not base:
        return prefix
    if base.startswith("["):
        return base
    return f"{prefix} {base}"


async def _require_customer_by_code(customer_code: str) -> Customer:
    code = (customer_code or "").strip()
    if not code:
        raise BadRequest("customer_code 不能为空")
    c = await Customer.filter(existed=True, code=code).first()
    if not c:
        raise BadRequest(
            f"客户不存在（客户编号={code}）。请先在【客户支持-客户】中创建客户并填写“客户编号(code)”。"
            f"注意这里的客户编号是客户的 code，不是客户ID。"
        )
    return c


async def _get_or_create_wallet(customer_code: str, customer_name: str | None = None) -> WalletAccount:
    await _require_customer_by_code(customer_code)
    w = await WalletAccount.filter(customer_code=customer_code).first()
    if w:
        if customer_name and w.customer_name != customer_name:
            w.customer_name = customer_name
            await w.save()
        return w
    return await WalletAccount.create(customer_code=customer_code, customer_name=customer_name)


async def _import_wallet_transactions_from_rows(
    rows: Iterable[list[str]],
    skip_header: bool,
    apply_to_balance: bool,
    allow_negative_balance: bool,
) -> ImportResult:
    """
    导入扣费流水（钱包流水：consume）。

    列顺序（推荐）：
    1) time: 扣费时间（如 2026-03-01 12:30:00）
    2) customer_code
    3) customer_name（可选）
    4) cost_type: stream/sms/token（或中文“流量/短信/token”）
    5) amount: 扣费金额（正数）
    6) remark（可选）
    7) related_bill_id（可选）
    8) related_order_id（可选）
    9) tx_id（可选；用于幂等，存在则跳过）
    """
    result = ImportResult()
    idx = 0
    for cols in rows:
        if skip_header and idx == 0:
            idx += 1
            continue
        idx += 1
        cols = [c.strip() for c in cols if c is not None]
        if len(cols) < 5:
            result.failed += 1
            continue

        (
            time_s,
            customer_code,
            customer_name,
            cost_type,
            amount,
            *rest,
        ) = cols

        remark = rest[0] if len(rest) >= 1 else None
        related_bill_id = rest[1] if len(rest) >= 2 else None
        related_order_id = rest[2] if len(rest) >= 3 else None
        tx_id = rest[3] if len(rest) >= 4 else None

        try:
            dt = arrow.get(time_s).naive
            amount_v = Decimal(str(amount))
            if amount_v <= 0:
                raise ValueError("amount must be > 0")
        except Exception:
            result.failed += 1
            continue

        try:
            wallet = await _get_or_create_wallet(customer_code=customer_code, customer_name=customer_name or None)
            if tx_id:
                exist = await WalletTransaction.filter(tx_id=tx_id).first()
                if exist:
                    result.skipped += 1
                    continue

            wallet.balance = Decimal(str(wallet.balance or 0))
            change_amount = -amount_v
            if apply_to_balance:
                target_balance = wallet.balance + change_amount
                if (not allow_negative_balance) and target_balance < 0:
                    raise BadRequest("导入会导致余额为负，已拒绝（可勾选允许负余额或关闭同步余额）")
                wallet.balance = target_balance
                await wallet.save()
                balance_after = wallet.balance
            else:
                balance_after = wallet.balance

            create_kwargs = dict(
                wallet=wallet,
                tx_type=WalletTransactionType.consume.value,
                change_amount=change_amount,
                balance_after=balance_after,
                related_order_id=related_order_id or None,
                related_bill_id=related_bill_id or None,
                remark=_build_cost_remark(cost_type, remark),
                create_time=dt,
            )
            if tx_id:
                create_kwargs["tx_id"] = tx_id
            await WalletTransaction.create(**create_kwargs)
            result.created += 1
        except Exception:
            result.failed += 1
            continue

    return result


@router.post(path_postfix="/orders", tags=[APITags.console], summary="导入订单（仅导入入库，不立即扣费）")
async def _(
    header: Annotated[HeaderToken, Header()],
    body: ImportTextBody,
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = [[c.strip() for c in raw.split("\t")] for raw in _lines(body.text)]
    result = await _import_orders_from_rows(rows, skip_header=body.skip_header)
    return JsonResp(data=result)


@router.post(path_postfix="/orders/file", tags=[APITags.console], summary="上传导入订单（CSV / Excel）")
async def _(
    header: Annotated[HeaderToken, Header()],
    file: UploadFile = File(...),
    skip_header: bool = Form(default=True),
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = await _iter_rows_from_upload(file)
    result = await _import_orders_from_rows(rows, skip_header=skip_header)
    return JsonResp(data=result)


@router.post(path_postfix="/month-bills", tags=[APITags.console], summary="导入月度账单（TSV）")
async def _(
    header: Annotated[HeaderToken, Header()],
    body: ImportTextBody,
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = [[c.strip() for c in raw.split("\t")] for raw in _lines(body.text)]
    result = await _import_month_bills_from_rows(rows, skip_header=body.skip_header)
    return JsonResp(data=result)


@router.post(path_postfix="/month-bills/file", tags=[APITags.console], summary="上传导入月度账单（CSV / Excel）")
async def _(
    header: Annotated[HeaderToken, Header()],
    file: UploadFile = File(...),
    skip_header: bool = Form(default=True),
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = await _iter_rows_from_upload(file)
    result = await _import_month_bills_from_rows(rows, skip_header=skip_header)
    return JsonResp(data=result)


@router.post(path_postfix="/wallet-transactions", tags=[APITags.console], summary="导入扣费记录（TSV 文本）")
async def _(
    header: Annotated[HeaderToken, Header()],
    body: ImportTextBody,
    apply_to_balance: bool = Query(default=False),
    allow_negative_balance: bool = Query(default=True),
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = [[c.strip() for c in raw.split("\t")] for raw in _lines(body.text)]
    result = await _import_wallet_transactions_from_rows(
        rows,
        skip_header=body.skip_header,
        apply_to_balance=apply_to_balance,
        allow_negative_balance=allow_negative_balance,
    )
    return JsonResp(data=result)


@router.post(path_postfix="/wallet-transactions/file", tags=[APITags.console], summary="上传导入扣费记录（CSV / Excel）")
async def _(
    header: Annotated[HeaderToken, Header()],
    file: UploadFile = File(...),
    skip_header: bool = Form(default=True),
    apply_to_balance: bool = Form(default=False),
    allow_negative_balance: bool = Form(default=True),
) -> JsonResp[ImportResult]:
    await header.verify(Privileges.system_control)
    rows = await _iter_rows_from_upload(file)
    result = await _import_wallet_transactions_from_rows(
        rows,
        skip_header=skip_header,
        apply_to_balance=apply_to_balance,
        allow_negative_balance=allow_negative_balance,
    )
    return JsonResp(data=result)

